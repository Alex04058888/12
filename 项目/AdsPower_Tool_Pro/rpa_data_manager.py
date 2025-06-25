#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
RPA数据管理系统
完全按照AdsPower原版实现数据管理，支持文件操作、Excel操作、数据转换等
"""

import json
import os
import csv
import random
import re
import openpyxl
from datetime import datetime
from typing import Any, Dict, List, Optional, Union

class RPADataManager:
    """RPA数据管理器 - 完全按照AdsPower原版实现"""
    
    def __init__(self):
        self.data_cache = {}  # 数据缓存
        self.file_cache = {}  # 文件缓存
        self.excel_cache = {}  # Excel缓存
        
    # ==================== 文件操作 ====================
    
    def save_to_file(self, file_path: str, content: str, encoding: str = "utf-8", mode: str = "w") -> bool:
        """保存内容到文件 - 按照AdsPower原版"""
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            with open(file_path, mode, encoding=encoding) as f:
                f.write(content)
            
            return True
        except Exception as e:
            print(f"保存文件失败: {str(e)}")
            return False
    
    def read_from_file(self, file_path: str, encoding: str = "utf-8") -> Optional[str]:
        """从文件读取内容 - 按照AdsPower原版"""
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return f.read()
        except Exception as e:
            print(f"读取文件失败: {str(e)}")
            return None
    
    def read_file_lines(self, file_path: str, encoding: str = "utf-8") -> List[str]:
        """按行读取文件 - 按照AdsPower原版"""
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                return [line.strip() for line in f.readlines()]
        except Exception as e:
            print(f"读取文件行失败: {str(e)}")
            return []
    
    def read_random_line(self, file_path: str, encoding: str = "utf-8") -> Optional[str]:
        """随机读取文件中的一行 - 按照AdsPower原版"""
        lines = self.read_file_lines(file_path, encoding)
        if lines:
            return random.choice(lines)
        return None
    
    # ==================== Excel操作 ====================
    
    def save_to_excel(self, file_path: str, data: List[List[Any]], sheet_name: str = "Sheet1", 
                     start_row: int = 1, start_col: int = 1) -> bool:
        """保存数据到Excel - 按照AdsPower原版"""
        try:
            # 如果文件存在，打开现有文件；否则创建新文件
            if os.path.exists(file_path):
                workbook = openpyxl.load_workbook(file_path)
            else:
                workbook = openpyxl.Workbook()
                # 删除默认的Sheet
                if "Sheet" in workbook.sheetnames:
                    workbook.remove(workbook["Sheet"])
            
            # 创建或获取工作表
            if sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.create_sheet(sheet_name)
            
            # 写入数据
            for row_idx, row_data in enumerate(data):
                for col_idx, cell_value in enumerate(row_data):
                    worksheet.cell(
                        row=start_row + row_idx,
                        column=start_col + col_idx,
                        value=cell_value
                    )
            
            # 保存文件
            workbook.save(file_path)
            return True
            
        except Exception as e:
            print(f"保存Excel失败: {str(e)}")
            return False
    
    def read_from_excel(self, file_path: str, sheet_name: str = None, 
                       start_row: int = 1, start_col: int = 1, 
                       max_rows: int = None) -> List[List[Any]]:
        """从Excel读取数据 - 按照AdsPower原版"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            
            # 选择工作表
            if sheet_name:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook.active
            
            data = []
            max_row = worksheet.max_row if max_rows is None else min(start_row + max_rows - 1, worksheet.max_row)
            
            for row in worksheet.iter_rows(min_row=start_row, max_row=max_row, 
                                         min_col=start_col, values_only=True):
                # 过滤空行
                if any(cell is not None for cell in row):
                    data.append(list(row))
            
            return data
            
        except Exception as e:
            print(f"读取Excel失败: {str(e)}")
            return []
    
    def get_excel_sheets(self, file_path: str) -> List[str]:
        """获取Excel文件的所有工作表名称"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            return workbook.sheetnames
        except Exception as e:
            print(f"获取Excel工作表失败: {str(e)}")
            return []
    
    # ==================== 数据转换 ====================
    
    def convert_to_json(self, data: Any, indent: int = None) -> str:
        """转换数据为JSON - 按照AdsPower原版"""
        try:
            return json.dumps(data, ensure_ascii=False, indent=indent)
        except Exception as e:
            print(f"转换JSON失败: {str(e)}")
            return "{}"
    
    def parse_json(self, json_str: str) -> Any:
        """解析JSON字符串 - 按照AdsPower原版"""
        try:
            return json.loads(json_str)
        except Exception as e:
            print(f"解析JSON失败: {str(e)}")
            return None
    
    def format_json(self, json_str: str, indent: int = 2) -> str:
        """格式化JSON字符串"""
        try:
            data = json.loads(json_str)
            return json.dumps(data, ensure_ascii=False, indent=indent)
        except Exception as e:
            print(f"格式化JSON失败: {str(e)}")
            return json_str
    
    def compress_json(self, json_str: str) -> str:
        """压缩JSON字符串"""
        try:
            data = json.loads(json_str)
            return json.dumps(data, ensure_ascii=False, separators=(',', ':'))
        except Exception as e:
            print(f"压缩JSON失败: {str(e)}")
            return json_str
    
    # ==================== 文本处理 ====================
    
    def extract_text_by_regex(self, text: str, pattern: str, extract_type: str = "first", 
                             case_sensitive: bool = True) -> Union[str, List[str], int]:
        """使用正则表达式提取文本 - 按照AdsPower原版"""
        try:
            flags = 0 if case_sensitive else re.IGNORECASE
            
            if extract_type == "all":
                matches = re.findall(pattern, text, flags)
                return matches
            elif extract_type == "first":
                match = re.search(pattern, text, flags)
                return match.group(0) if match else ""
            elif extract_type == "last":
                matches = re.findall(pattern, text, flags)
                return matches[-1] if matches else ""
            elif extract_type == "count":
                matches = re.findall(pattern, text, flags)
                return len(matches)
            else:
                match = re.search(pattern, text, flags)
                return match.group(0) if match else ""
                
        except Exception as e:
            print(f"正则提取失败: {str(e)}")
            return "" if extract_type != "count" else 0
    
    def extract_field_from_data(self, data: Any, field_path: str, default_value: Any = None) -> Any:
        """从复杂数据结构中提取字段 - 按照AdsPower原版"""
        try:
            current = data
            
            # 解析字段路径，支持 user.name 和 data[0].title 格式
            parts = re.split(r'[\.\[\]]', field_path)
            parts = [part for part in parts if part]  # 过滤空字符串
            
            for part in parts:
                if part.isdigit():
                    # 数组索引
                    current = current[int(part)]
                else:
                    # 对象属性
                    if isinstance(current, dict):
                        current = current.get(part)
                    else:
                        current = getattr(current, part, None)
                
                if current is None:
                    return default_value
            
            return current
            
        except Exception as e:
            print(f"字段提取失败: {str(e)}")
            return default_value
    
    def random_extract_from_data(self, data: Any, extract_type: str = "random_line", 
                                count: int = 1, unique: bool = True) -> Any:
        """从数据中随机提取 - 按照AdsPower原版"""
        try:
            items = []
            
            if extract_type == "random_line":
                if isinstance(data, str):
                    items = [line.strip() for line in data.split('\n') if line.strip()]
                elif isinstance(data, list):
                    items = [str(item) for item in data]
                else:
                    items = [str(data)]
            elif extract_type == "random_element":
                if isinstance(data, (list, tuple)):
                    items = list(data)
                elif isinstance(data, dict):
                    items = list(data.values())
                else:
                    items = [data]
            elif extract_type == "random_char":
                if isinstance(data, str):
                    items = list(data)
                else:
                    items = list(str(data))
            elif extract_type == "random_word":
                if isinstance(data, str):
                    items = re.findall(r'\b\w+\b', data)
                else:
                    items = re.findall(r'\b\w+\b', str(data))
            
            if not items:
                return "" if count == 1 else []
            
            # 执行随机提取
            if unique and count >= len(items):
                result = items.copy()
                random.shuffle(result)
            elif unique:
                result = random.sample(items, count)
            else:
                result = [random.choice(items) for _ in range(count)]
            
            return result[0] if count == 1 else result
            
        except Exception as e:
            print(f"随机提取失败: {str(e)}")
            return "" if count == 1 else []
    
    # ==================== 数据缓存 ====================
    
    def cache_data(self, key: str, data: Any, ttl: int = 3600):
        """缓存数据"""
        self.data_cache[key] = {
            "data": data,
            "timestamp": datetime.now().timestamp(),
            "ttl": ttl
        }
    
    def get_cached_data(self, key: str) -> Any:
        """获取缓存数据"""
        if key in self.data_cache:
            cache_item = self.data_cache[key]
            if datetime.now().timestamp() - cache_item["timestamp"] < cache_item["ttl"]:
                return cache_item["data"]
            else:
                del self.data_cache[key]
        return None
    
    def clear_cache(self):
        """清除所有缓存"""
        self.data_cache.clear()
        self.file_cache.clear()
        self.excel_cache.clear()
    
    # ==================== 数据验证 ====================
    
    def validate_json(self, json_str: str) -> bool:
        """验证JSON格式"""
        try:
            json.loads(json_str)
            return True
        except:
            return False
    
    def validate_file_exists(self, file_path: str) -> bool:
        """验证文件是否存在"""
        return os.path.exists(file_path)
    
    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """获取文件信息"""
        try:
            stat = os.stat(file_path)
            return {
                "exists": True,
                "size": stat.st_size,
                "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(),
                "created": datetime.fromtimestamp(stat.st_ctime).isoformat()
            }
        except:
            return {"exists": False}
    
    def get_data_info(self) -> Dict[str, Any]:
        """获取数据管理器信息"""
        return {
            "cached_items": len(self.data_cache),
            "file_cache_items": len(self.file_cache),
            "excel_cache_items": len(self.excel_cache),
            "last_update": datetime.now().isoformat()
        }
